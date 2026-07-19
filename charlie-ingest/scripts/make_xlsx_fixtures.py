# Concern: AUTHOR the committed .xlsx test corpus for charlie-ingest — emit a small set of real Excel (OOXML) spreadsheets (via openpyxl) MIRRORING the .ods corpus (every value type, a SUM/formula chain, a cross-sheet reference, date cells, blanks, VLOOKUP/IF) PLUS an xlsx-only reference-resolution fixture (`resolution.xlsx`: a worksheet TABLE with structured references `Sales[Q1]`/`Sales[@Q1]`/`Sales[[#Headers],[Q1]]` and workbook DEFINED NAMES `TaxRate`/`AllQ1`, with hand-verified expected values) — so the same import->eval assertions hold and the importer's name/table -> A1 resolution is proven end-to-end. An xlsx formula is stored in Excel-A1 already (no `of:=`/`[.A1]`), and an openpyxl-written formula carries NO cached value (charlie re-evaluates), so these files exercise the reader's format-blind path, translate's near-noop-for-xlsx behaviour, and the resolver | Non-concern: the Rust importer/translation under test (src/**) and running the tests (the .xlsx artifacts this writes are committed; the venv that runs this is gitignored); the .ods corpus (make_fixtures.py owns that) | IO: () -> writes tests/fixtures/*.xlsx next to this script's parent crate
"""Regenerate the committed .xlsx fixture corpus. Run from a venv with openpyxl installed:

    python3 -m venv .fixture-venv && .fixture-venv/bin/pip install openpyxl
    .fixture-venv/bin/python scripts/make_xlsx_fixtures.py

The generated .xlsx files ARE committed (small), so the Rust tests need no python/openpyxl. These
mirror the .ods fixtures (make_fixtures.py) cell-for-cell so tests/import_xlsx.rs and the .ods
integration test assert the same Excel-correct values through the one format-blind engine.
"""

import datetime
import os

import openpyxl

FIXTURES = os.path.join(os.path.dirname(__file__), "..", "tests", "fixtures")


def write(wb, name):
    os.makedirs(FIXTURES, exist_ok=True)
    path = os.path.join(FIXTURES, name)
    wb.save(path)
    print("wrote", os.path.normpath(path))


def make_smoke():
    """The batch smoke + cross-sheet fixture:
    Sheet1: A1=10, A2=20, A3=SUM(A1:A2)=30, B1=A3*2=60
    Sheet2: A1==Sheet1!A3=30
    """
    wb = openpyxl.Workbook()
    s1 = wb.active
    s1.title = "Sheet1"
    s1["A1"] = 10
    s1["A2"] = 20
    s1["A3"] = "=SUM(A1:A2)"
    s1["B1"] = "=A3*2"
    s2 = wb.create_sheet("Sheet2")
    s2["A1"] = "=Sheet1!A3"
    write(wb, "smoke.xlsx")


def make_literals():
    """Every literal value type + date cells + an interior/trailing blank."""
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Data"
    for col, head in enumerate(["Name", "Qty", "Active", "When"], start=1):
        t.cell(row=1, column=col, value=head)
    # Widget row: text, number, bool, date. 2024-01-15 -> Excel serial 45306.
    t["A2"] = "Widget"
    t["B2"] = 42
    t["C2"] = True
    t["D2"] = datetime.datetime(2024, 1, 15)
    # Gadget row: a negative float, an interior blank (no Active), another date (2024-06-30 -> 45473).
    t["A3"] = "Gadget"
    t["B3"] = -3.5
    t["D3"] = datetime.datetime(2024, 6, 30)
    write(wb, "literals.xlsx")


def make_functions():
    """VLOOKUP + IF over a small table (Excel-A1 args: ',' separators, quoted string, bare FALSE)."""
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Lookup"
    # Columns A,B are the table; column D holds the function cells (C left blank).
    t["A1"] = "apple"
    t["B1"] = 1
    t["D1"] = '=VLOOKUP("banana",A1:B3,2,FALSE)'
    t["A2"] = "banana"
    t["B2"] = 2
    t["D2"] = '=IF(B1>0,"pos","neg")'
    t["A3"] = "cherry"
    t["B3"] = 3
    write(wb, "functions.xlsx")


def make_resolution():
    """Defined names + table structured references (the import-time reference-resolution fixture).

    Sheet "Data" holds a table `Sales` (A1:C4; header row Region/Q1/Q2; data rows 2..4) plus formulas
    that use structured refs and defined names, with hand-verified expected values:
        E2 =SUM(Sales[Q1])           -> SUM(B2:B4) = 10+20+30 = 60
        E3 =SUM(Sales[Q2])           -> SUM(C2:C4) = 15+25+35 = 75
        F2 =Sales[@Q1]               -> B2 = 10   (this-row, formula on row 2)
        F3 =Sales[@Q1]               -> B3 = 20   (this-row, formula on row 3)
        G2 =Sales[[#Headers],[Q1]]   -> B1 = "Q1" (the header cell)
        H2 =TaxRate*100              -> Data!$H$1 * 100 = 0.2*100 = 20
        H3 =SUM(AllQ1)               -> SUM(Data!$B$2:$B$4) = 60
    Defined names: TaxRate -> Data!$H$1 (a cell), AllQ1 -> Data!$B$2:$B$4 (a range). openpyxl writes no
    cached values, so charlie re-evaluates every formula from the resolved A1 the importer produced.
    """
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.table import Table

    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Data"
    # The Sales table: a header row + three data rows.
    for col, head in enumerate(["Region", "Q1", "Q2"], start=1):
        t.cell(row=1, column=col, value=head)
    for r, (region, q1, q2) in enumerate(
        [("North", 10, 15), ("South", 20, 25), ("East", 30, 35)], start=2
    ):
        t.cell(row=r, column=1, value=region)
        t.cell(row=r, column=2, value=q1)
        t.cell(row=r, column=3, value=q2)
    t.add_table(Table(displayName="Sales", ref="A1:C4"))

    # Structured references (columns E/F/G) + defined-name references (column H).
    t["E2"] = "=SUM(Sales[Q1])"
    t["E3"] = "=SUM(Sales[Q2])"
    t["F2"] = "=Sales[@Q1]"
    t["F3"] = "=Sales[@Q1]"
    t["G2"] = "=Sales[[#Headers],[Q1]]"
    t["H1"] = 0.2
    t["H2"] = "=TaxRate*100"
    t["H3"] = "=SUM(AllQ1)"
    wb.defined_names.add(DefinedName("TaxRate", attr_text="Data!$H$1"))
    wb.defined_names.add(DefinedName("AllQ1", attr_text="Data!$B$2:$B$4"))
    write(wb, "resolution.xlsx")


def make_blanks_repeats():
    """A sparse sheet: leading value, interior blanks, a trailing value, a blank row. Used range A1:D3."""
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Sparse"
    # A1=1, B1/C1 blank, D1=4; row 2 fully blank; A3=7.
    t["A1"] = 1
    t["D1"] = 4
    t["A3"] = 7
    write(wb, "blanks_repeats.xlsx")


if __name__ == "__main__":
    make_smoke()
    make_literals()
    make_functions()
    make_blanks_repeats()
    make_resolution()
