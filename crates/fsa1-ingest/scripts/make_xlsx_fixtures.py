# Concern: authors the committed .xlsx fixture corpus | Non-concern: the Rust importer, the .ods corpus | IO: () -> tests/fixtures/*.xlsx
"""Run from a venv with openpyxl installed; the generated .xlsx files are committed, so the Rust
tests need neither. These mirror make_fixtures.py cell-for-cell, so both formats assert the same
Excel-correct values through the one format-blind engine.

    python3 -m venv .fixture-venv && .fixture-venv/bin/pip install openpyxl
    .fixture-venv/bin/python scripts/make_xlsx_fixtures.py
"""

import datetime
import os

import openpyxl

FIXTURES = os.path.join(os.path.dirname(__file__), "..", "tests", "fixtures")


def write(wb, name):
    os.makedirs(FIXTURES, exist_ok=True)
    path = os.path.join(FIXTURES, name)
    wb.save(path)
    print("wrote", os.path.normpath(path))


def make_smoke():
    wb = openpyxl.Workbook()
    s1 = wb.active
    s1.title = "Sheet1"
    s1["A1"] = 10
    s1["A2"] = 20
    s1["A3"] = "=SUM(A1:A2)"
    s1["B1"] = "=A3*2"
    s2 = wb.create_sheet("Sheet2")
    s2["A1"] = "=Sheet1!A3"
    write(wb, "smoke.xlsx")


def make_literals():
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Data"
    for col, head in enumerate(["Name", "Qty", "Active", "When"], start=1):
        t.cell(row=1, column=col, value=head)
    # 2024-01-15 is Excel serial 45306; 2024-06-30 is 45473.
    t["A2"] = "Widget"
    t["B2"] = 42
    t["C2"] = True
    t["D2"] = datetime.datetime(2024, 1, 15)
    t["A3"] = "Gadget"
    t["B3"] = -3.5
    t["D3"] = datetime.datetime(2024, 6, 30)
    write(wb, "literals.xlsx")


def make_functions():
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Lookup"
    # Columns A,B are the table; column D holds the function cells (C left blank).
    t["A1"] = "apple"
    t["B1"] = 1
    t["D1"] = '=VLOOKUP("banana",A1:B3,2,FALSE)'
    t["A2"] = "banana"
    t["B2"] = 2
    t["D2"] = '=IF(B1>0,"pos","neg")'
    t["A3"] = "cherry"
    t["B3"] = 3
    write(wb, "functions.xlsx")


def make_future_functions():
    """Excel STORES a post-2007 function as `_xlfn.NAME` while DISPLAYING the bare name, and
    openpyxl stores a formula string uninterpreted, so writing the prefixed text here reproduces
    the real on-disk shape. C3 is the unprefixed control.
    """
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Future"
    s["A1"] = 3
    s["A2"] = 1
    s["A3"] = 2
    s["C1"] = '=_xlfn.MINIFS(A1:A3,A1:A3,">1")'
    s["C2"] = "=_xlfn.XLOOKUP(2,A1:A3,A1:A3)"
    s["C3"] = "=SUM(A1:A3)"
    write(wb, "future_functions.xlsx")


def make_resolution():
    """The name is `AllQOne`, not `AllQ1`, because an identifier that parses as an A1 address is
    skipped at emit. openpyxl writes no cached values, so FSA1 recomputes every formula from
    the resolved A1 the importer produced.
    """
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.table import Table

    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Data"
    for col, head in enumerate(["Region", "Q1", "Q2"], start=1):
        t.cell(row=1, column=col, value=head)
    for r, (region, q1, q2) in enumerate(
        [("North", 10, 15), ("South", 20, 25), ("East", 30, 35)], start=2
    ):
        t.cell(row=r, column=1, value=region)
        t.cell(row=r, column=2, value=q1)
        t.cell(row=r, column=3, value=q2)
    t.add_table(Table(displayName="Sales", ref="A1:C4"))

    t["E2"] = "=SUM(Sales[Q1])"
    t["E3"] = "=SUM(Sales[Q2])"
    t["F2"] = "=Sales[@Q1]"
    t["F3"] = "=Sales[@Q1]"
    t["G2"] = "=Sales[[#Headers],[Q1]]"
    t["H1"] = 0.2
    t["H2"] = "=TaxRate*100"
    t["H3"] = "=SUM(AllQOne)"
    wb.defined_names.add(DefinedName("TaxRate", attr_text="Data!$H$1"))
    wb.defined_names.add(DefinedName("AllQOne", attr_text="Data!$B$2:$B$4"))
    write(wb, "resolution.xlsx")


def make_fidelity():
    """Passes the strict pre-flight (all General, no tail parts) yet still loses two things: a
    defined name whose identifier is an A1 address, and an untranslatable inline-array formula.
    """
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Data"
    t["A1"] = 5
    t["B1"] = "=SUM({1,2,3})"
    t["C1"] = 7
    wb.defined_names.add(DefinedName("A1", attr_text="Data!$C$1"))
    write(wb, "fidelity.xlsx")


def make_many_formats():
    """Sixty coerced cells, so the fidelity report is exercised well past any truncation cap."""
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Nums"
    for r in range(1, 61):
        c = t.cell(row=r, column=1, value=r * 10)
        c.number_format = '"$"#,##0.00'
    write(wb, "many_formats.xlsx")


def make_visuals():
    """Everything the read side must recover beyond a value: a fully styled cell, a cell whose ONLY
    content is its fill (openpyxl writes it as `<c r="B2" s="n"/>`, with no `<v>`, which is the shape
    calamine never yields), a custom column width and row height, and a merged region.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.styles.colors import Color

    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Visual"

    edge = Side(style="thin", color="FFFF0000")
    a1 = t["A1"]
    a1.value = "Header"
    a1.font = Font(
        name="Times New Roman",
        size=14,
        bold=True,
        italic=True,
        underline="single",
        strike=True,
        color=Color(theme=4, tint=0.4),
    )
    a1.fill = PatternFill(fill_type="solid", fgColor="FFFFC000")
    a1.border = Border(left=edge, right=edge, top=edge, bottom=edge)
    a1.alignment = Alignment(horizontal="center", vertical="top", wrapText=True, indent=2)

    t["B2"].fill = PatternFill(fill_type="solid", fgColor="FF00B0F0")
    t["A4"] = 5
    t.column_dimensions["C"].width = 14.5
    t.row_dimensions[3].height = 22.5
    t.merge_cells("D1:E1")
    write(wb, "visuals.xlsx")


def make_gapped_columns():
    """Two column regions with one empty column between them, and a width authored ON that column.
    Nothing states a style, so both rectangles cost zero rules and bridging column C gains nothing:
    the default cut leaves C in no block, and the width has no range file to ride on.
    """
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Gapped"
    for r in range(1, 21):
        for c in (1, 2, 4, 5):
            t.cell(row=r, column=c, value=r * 10 + c)
    t.column_dimensions["C"].width = 14.5
    write(wb, "gapped_columns.xlsx")


def make_blanks_repeats():
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Sparse"
    # Used range A1:D3: A1=1, B1/C1 blank, D1=4; row 2 fully blank; A3=7.
    t["A1"] = 1
    t["D1"] = 4
    t["A3"] = 7
    write(wb, "blanks_repeats.xlsx")


if __name__ == "__main__":
    make_smoke()
    make_literals()
    make_functions()
    make_future_functions()
    make_blanks_repeats()
    make_gapped_columns()
    make_resolution()
    make_fidelity()
    make_many_formats()
    make_visuals()
