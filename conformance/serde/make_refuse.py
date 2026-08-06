# Concern: authors the refuse/ probes (SER2/SER3) — an out-of-scope .xlsx per --strict refusal trigger | Non-concern: computing or diffing values, the accepted fixtures | IO: () -> refuse/*.xlsx
"""Regenerate the SYNTHESIZED refuse/ probes for the serde round-trip conformance corpus (plan 07 §8).

Run via the oracle venv (created by run.sh):

    conformance/serde/.venv/bin/python conformance/serde/make_refuse.py

The generated .xlsx ARE committed (small, LFS-tracked); the venv is gitignored. Provenance is honest
by construction: openpyxl (a third-party writer) authors the bytes, never FSA1. Each probe carries
exactly one part/cell OUTSIDE the skeleton's ALLOW-set (or the accepted numFmt catalog) so `fsa1-cli
unpack --strict` refuses it with a located CORE2 diagnostic (SER3). resolution.xlsx (an xl/tables part)
is graduated ready-made from fsa1-ingest/tests/fixtures/ and is NOT re-authored here.

GRID7 note (plan 07): the pre-amendment ``numfmt.xlsx`` (a `0.00%` cell) and ``literals.xlsx`` (a
datetime cell) are RETIRED as refuse probes — percent is an ACCEPTED category and a formatted formula
with an accepted format now ACCEPTS. The refuse set here is the genuine EXOTIC TAIL: value-dependent
conditional switches and digit masks the self-describing content cannot carry, on a literal AND a
formula alike.
"""

import os
import shutil
import zipfile

import openpyxl
from openpyxl.chart import BarChart, Reference

HERE = os.path.dirname(__file__)
OUT = os.path.join(HERE, "refuse")


def _save(wb, name):
    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, name)
    wb.save(path)
    return path


def make_cond_literal():
    """A value literal under a CONDITIONAL-SWITCH numFmt (`[<100]0.00;[>=100]0`) — a value-dependent
    format the self-describing content cannot carry (the exotic tail, §4.1). `unpack --strict` names the
    cell + numFmtId + formatCode. The value (50) is otherwise unremarkable; the FORMAT alone refuses."""
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Exotic"
    s["A1"] = 50
    s["A1"].number_format = "[<100]0.00;[>=100]0"
    _save(wb, "cond_literal.xlsx")
    print("wrote refuse/cond_literal.xlsx (conditional-switch numFmt on Exotic!A1)")


def make_mask_literal():
    """A value literal under a DIGIT/PHONE MASK (`000000000`) — a leading-zero mask better modeled as
    TEXT than a formatted Number (the exotic tail, §4.1). `unpack --strict` names the cell + numFmtId."""
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Exotic"
    s["A1"] = 123456789
    s["A1"].number_format = "000000000"
    _save(wb, "mask_literal.xlsx")
    print("wrote refuse/mask_literal.xlsx (digit mask numFmt on Exotic!A1)")


def make_exotic_formula():
    """A FORMULA cell under a CONDITIONAL-SWITCH numFmt — proving the exotic tail refuses BOTH a literal
    AND a formula (the format is exotic, a catalog concern, not the literal-only precision concern). The
    formula computes fine; the FORMAT is unrepresentable, so `unpack --strict` refuses naming the cell."""
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Exotic"
    s["A1"] = 5
    s["A2"] = "=A1*2"
    s["A2"].number_format = "[<100]0;[>=100]0.0"
    _save(wb, "exotic_formula.xlsx")
    print("wrote refuse/exotic_formula.xlsx (conditional-switch numFmt on a FORMULA cell Exotic!A2)")


def make_chart():
    """A workbook carrying an embedded chart -> openpyxl emits xl/charts/chart1.xml (+ a drawing). A
    chart is a part FSA1 neither models nor can regenerate (SER3 REFUSE role); its presence is the
    refusal trigger."""
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Data"
    for i, v in enumerate([3, 7, 2, 9], start=1):
        s[f"A{i}"] = v
    chart = BarChart()
    chart.add_data(Reference(s, min_col=1, min_row=1, max_row=4))
    s.add_chart(chart, "C1")
    _save(wb, "chart.xlsx")
    print("wrote refuse/chart.xlsx (embedded chart -> xl/charts/chart1.xml)")


# A minimal, well-formed OOXML drawing part — enough to be a real out-of-ALLOW package part. FSA1
# refuses on its PRESENCE (SER3), so it need not be a fully-wired drawingML anchor.
_DRAWING_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
    ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"/>'
)

# A minimal pivotTable part. openpyxl does not author pivot parts natively, so it is injected into the
# package from this script (the plan sanctions this) — again, PRESENCE is what SER3 refuses on.
_PIVOT_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
    ' name="PivotProbe" cacheId="0" dataOnRows="1" applyNumberFormats="0" applyBorderFormats="0"'
    ' applyFontFormats="0" applyPatternFormats="0" applyAlignmentFormats="0"'
    ' applyWidthHeightFormats="1" dataCaption="Values"/>'
)


def _inject_part(name, part_path, part_xml):
    """Author a plain single-cell workbook with openpyxl, then rewrite the package to INSERT one extra
    raw part (`part_path`) — an out-of-ALLOW part whose mere presence trips SER3. Rewriting the zip
    (rather than openpyxl) is how a part openpyxl cannot author natively is introduced honestly."""
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Probe"
    s["A1"] = 1
    base = _save(wb, name)
    tmp = base + ".tmp"
    with zipfile.ZipFile(base, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            zout.writestr(item, zin.read(item.filename))
        zout.writestr(part_path, part_xml)
    shutil.move(tmp, base)
    print(f"wrote refuse/{name} (injected out-of-ALLOW part {part_path})")


def make_drawing():
    _inject_part("drawing.xlsx", "xl/drawings/drawing1.xml", _DRAWING_XML)


def make_pivot():
    _inject_part("pivot.xlsx", "xl/pivotTables/pivotTable1.xml", _PIVOT_XML)


if __name__ == "__main__":
    make_cond_literal()
    make_mask_literal()
    make_exotic_formula()
    make_chart()
    make_drawing()
    make_pivot()
