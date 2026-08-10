# Concern: authors the presentation corpus's .xlsx fixtures — one visual concern per file | Non-concern: unpacking, grading, the frozen expectations | IO: () -> fixtures/*.xlsx

"""Regenerate the presentation-conformance fixtures (plan 01 step 12).

Run via the oracle venv (created by run.sh):

    conformance/presentation/.venv/bin/python conformance/presentation/make_fixtures.py

The generated .xlsx ARE committed (small, LFS-tracked); the venv is gitignored. Provenance is honest
by construction: openpyxl (a third-party writer) authors every byte, never FSA1 — see PROVENANCE.md.

Each fixture isolates ONE concern and is named for it, so a grader failure names the concern rather
than "the corpus". Three families:

  * a STYLE fixture asserts what `unpack` must emit as the range file's trailing ``@scope`` block, and
    that packing it back preserves that block;
  * a ``warn_*`` fixture asserts a single SER3 warning and the ABSENCE of the ``nothing lost`` line;
  * a ``chart_*`` fixture asserts the ``.json`` figure a chart crosses as, or the named loss it
    costs where none can. ``openpyxl.chart`` authors every one, as openpyxl authors every other byte.

Anchors shared with the expectations, so a reviewer can cross-read the two: the Normal font of an
openpyxl workbook is Calibri 11pt unless the fixture says otherwise, and ``theme=1, tint=0`` is the
document's own default text colour (dk1) — a visual no-op that must produce NO ``color`` declaration.

EVERY fill colour is spelled 8-digit ARGB. openpyxl stores a 6-digit value with a ``00`` alpha, and a
pack writes ``FF`` unconditionally, so a 6-digit fill authors a colour no export can reproduce and the
reopen leg reads it as a divergence.
"""

import os

import openpyxl
from openpyxl.chart import BarChart, LineChart, RadarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "fixtures")

HEADER_FILL = "FFD9E1F2"
BLANK_FILL = "FF00B0F0"
BAND_FILL = "FFF2F2F2"
HATCH_PATTERN = "gray125"
BLACK = "FF000000"
COLUMN_WIDTH = 14.5
ROW_HEIGHT = 22.5
TITLE_COLUMN_WIDTH = 18.5
NORMAL_FONT_NAME = "Arial"
NORMAL_FONT_SIZE = 9


def _save(wb, name):
    os.makedirs(OUT, exist_ok=True)
    wb.save(os.path.join(OUT, name))
    print(f"wrote fixtures/{name}")


def _wb(title):
    wb = openpyxl.Workbook()
    wb.active.title = title
    return wb, wb.active


def make_styled_header_row():
    """A header row over a body: bold, filled and centred, uniform across row 1, so the encoder must
    reach for ONE `fsa1-row:first-child fsa1-cell` rule rather than three per-cell ones."""
    wb, s = _wb("Report")
    for col, text in zip("ABC", ("Region", "Units", "Revenue")):
        cell = s[f"{col}1"]
        cell.value = text
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center")
    s["A2"], s["B2"], s["C2"] = "North", 12, 3400
    s["A3"], s["B3"], s["C3"] = "South", 9, 2750
    _save(wb, "styled_header_row.xlsx")


def make_formatted_column():
    """A column uniform in a look its neighbours do not share — the Col leg of the cascade."""
    wb, s = _wb("Ledger")
    for row, (label, amount) in enumerate(
        [("Opening", 100), ("Fees", 25), ("Interest", 7), ("Closing", 82)], start=1
    ):
        s[f"A{row}"] = label
        cell = s[f"B{row}"]
        cell.value = amount
        cell.font = Font(name="Calibri", size=11, italic=True)
        cell.alignment = Alignment(horizontal="right")
    _save(wb, "formatted_column.xlsx")


def make_banded_body():
    """A zebra body whose shaded lines are EXACTLY the even ones — the periodic leg of the cascade."""
    wb, s = _wb("Report")
    shade = PatternFill(fill_type="solid", fgColor=BAND_FILL)
    for row, (label, amount) in enumerate(
        [("North", 10), ("South", 20), ("East", 30), ("West", 40), ("Inland", 50), ("Coast", 60)],
        start=1,
    ):
        s[f"A{row}"] = label
        s[f"B{row}"] = amount
        if row % 2 == 0:
            s[f"A{row}"].fill = shade
            s[f"B{row}"].fill = shade
    _save(wb, "banded_body.xlsx")


def make_total_row_top_border():
    """A total row ruled off from the body above it — a border on ONE edge of a whole row."""
    wb, s = _wb("Totals")
    for row, (label, amount) in enumerate([("Rent", 1500), ("Power", 210), ("Water", 90)], start=1):
        s[f"A{row}"] = label
        s[f"B{row}"] = amount
    s["A4"], s["B4"] = "Total", 1800
    top = Side(style="thin", color=BLACK)
    for col in "AB":
        cell = s[f"{col}4"]
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.border = Border(top=top)
    _save(wb, "total_row_top_border.xlsx")


def make_body_8pt_vs_normal_11pt():
    """Every cell at 8pt against the workbook's 11pt Normal: the size must cross as ONE `fsa1-cell` rule, and
    the Normal-equal family (Calibri) must NOT be emitted beside it."""
    wb, s = _wb("Body")
    for row in (1, 2, 3):
        for col in "AB":
            cell = s[f"{col}{row}"]
            cell.value = f"{col}{row}"
            cell.font = Font(name="Calibri", size=8)
    _save(wb, "body_8pt_vs_normal_11pt.xlsx")


def make_normal_font_arial_9():
    """The one fixture whose Normal cell style is NOT Calibri 11. Every other fixture inherits
    openpyxl's default Normal font, which happens to equal the format's own default — so the corpus
    could not see a workbook whose own default differs. Here the cells wear the Normal style and
    nothing else, so Arial 9 crosses only if the encoder declares it against the FORMAT's default;
    measuring against the SOURCE's Normal font drops the typeface and still prints `nothing lost`.

    The cells are stamped `Normal` rather than left bare so the fact is stated where openpyxl's reader
    looks: it resolves a cell's font through `cellXfs` alone and never follows `xfId` to the named
    style, so on a bare cell the reopen leg would compare against Calibri 11 whatever the workbook's
    Normal style says."""
    wb, s = _wb("Arial")
    wb._named_styles["Normal"].font = Font(name=NORMAL_FONT_NAME, size=NORMAL_FONT_SIZE)
    for row in (1, 2):
        for col in "AB":
            s[f"{col}{row}"] = f"{col}{row}"
            s[f"{col}{row}"].style = "Normal"
    _save(wb, "normal_font_arial_9.xlsx")


def make_sparse_blocks_normal_font_arial_9():
    """A non-default Normal font over SPARSE occupancy, so each block encloses blanks the source never
    stated. `normal_font_arial_9` is 2x2 and fully occupied, and `style_only_blank`'s blank carries a
    fill, so neither corpus entry holds the one shape that matters here: a blank whose only look is a
    typeface. A typeface needs a glyph to show, so such a blank is NOT occupancy — a leg that counts it
    as one turns every blank inside the block into content, and the two blocks fuse into one on each
    re-pack while the run still prints `nothing lost`.

    The four coordinates are sized against the partition's own waste budget, which is what makes the
    fusion visible: 5x60 = 300 cells over the 4 the source occupies is past the budget, so the 50-row
    gap cuts it into two 5x5 blocks; over the 50 a leg counting blanks would occupy, 300 is UNDER it,
    so nothing cuts and the second unpack writes one 300-field file."""
    wb, s = _wb("Gaps")
    wb._named_styles["Normal"].font = Font(name=NORMAL_FONT_NAME, size=NORMAL_FONT_SIZE)
    for coord, value in (("A1", 1), ("E5", 2), ("A56", 3), ("E60", 4)):
        s[coord] = value
        s[coord].style = "Normal"
    _save(wb, "sparse_blocks_normal_font_arial_9.xlsx")


def make_column_and_row_default_style():
    """A whole-column and a whole-row format, stated on the AXIS and on no cell. Excel and openpyxl
    both write these as `<col style=>` and `<row s= customFormat=>`, which is how "make this column
    bold" is encoded — and no other fixture holds either, so the corpus could not see a reader that
    took `s=` off `<c>` alone. Such a reader loses both looks entirely and still prints `nothing lost`.

    Two sheets so the two axes cannot interfere: a row's style replaces a column's on the cell where
    they cross, and one sheet holding both would grade that precedence instead of the carrying. Each
    is fully occupied, so what the statement does to the blanks it also dresses stays out of the
    reading here (`crates/fsa1-ingest` owns that half).

    openpyxl emits a `<col>` element only with its default width beside it, so column B's run states
    `width="13" customWidth="1"` as well; the width is incidental here — `axis_width_and_height` is
    what grades one — but it does show both facts being read off the ONE element."""
    wb, s = _wb("Column")
    for row in (1, 2):
        for col in "AB":
            s[f"{col}{row}"] = f"{col}{row}"
    s.column_dimensions["B"].font = Font(name="Calibri", size=11, bold=True)
    r = wb.create_sheet("Row")
    for row in (1, 2):
        for col in "AB":
            r[f"{col}{row}"] = f"{col}{row}"
    r.row_dimensions[2].font = Font(name="Calibri", size=11, italic=True)
    _save(wb, "column_and_row_default_style.xlsx")


def make_column_run_default_style():
    """A `<col style=>` run spanning SEVERAL columns, which `column_and_row_default_style` does not:
    its run is one column wide, so a reader that expands a run into per-cell styles and a reader that
    keeps it a run spell that fixture identically. Three columns cannot be told apart that way — the
    run either crosses as a column rule or it comes back as one rule per cell.

    The body is four rows so the per-cell spelling would be twelve rules against three, and the
    columns carry no cell-level style at all, so nothing but the axis statement can put the look
    back."""
    wb, s = _wb("Run")
    for row in range(1, 5):
        for col in "ABCD":
            s[f"{col}{row}"] = f"{col}{row}"
    for col in "BCD":
        s.column_dimensions[col].font = Font(name="Calibri", size=11, bold=True)
    _save(wb, "column_run_default_style.xlsx")


def make_theme1_color_noop():
    """Every cell carries `theme=1, tint=0` — the document's own default text colour. A visual no-op
    that must produce NO `color` declaration, while the bold beside it still crosses."""
    wb, s = _wb("Theme")
    for row in (1, 2):
        for col in "AB":
            cell = s[f"{col}{row}"]
            cell.value = f"{col}{row}"
            cell.font = Font(name="Calibri", size=11, bold=True, color=Color(theme=1, tint=0.0))
    _save(wb, "theme1_color_noop.xlsx")


def make_axis_width_and_height():
    """A non-default column width and row height, which must appear on disk as the .xlsx's own numbers
    — `width: 14.5ch` and `height: 22.5pt` — never a re-derived approximation."""
    wb, s = _wb("Axes")
    for row in (1, 2):
        for col in "AB":
            s[f"{col}{row}"] = f"{col}{row}"
    s.column_dimensions["A"].width = COLUMN_WIDTH
    s.row_dimensions[2].height = ROW_HEIGHT
    _save(wb, "axis_width_and_height.xlsx")


def make_title_over_table():
    """A title in A1:A2, an empty row 3, and a 4x17 table from row 4 — occupancy the budget rule keeps
    as ONE block, so the sheet writes ONE file and column A's width still has a container."""
    wb, s = _wb("Sheet1")
    s["A1"] = "Quarterly Report"
    s["A2"] = "FY2026"
    for col, head in zip("ABCD", ("Item", "Qty", "Price", "Total")):
        s[f"{col}4"] = head
    for row in range(5, 21):
        n = row - 4
        s[f"A{row}"] = f"Item {n}"
        s[f"B{row}"] = n
        s[f"C{row}"] = n * 3
        s[f"D{row}"] = n * n * 3
    s.column_dimensions["A"].width = TITLE_COLUMN_WIDTH
    _save(wb, "title_over_table.xlsx")


def make_stray_cell_sheet():
    """`A1`, `B1` and a stray `Z50000`. The stray must cost its own cell, not stretch a block across
    50,000 rows: TWO files totalling THREE TSV fields."""
    wb, s = _wb("Sparse")
    s["A1"] = 1
    s["B1"] = 2
    s["Z50000"] = 3
    _save(wb, "stray_cell_sheet.xlsx")


def make_single_cell_sheet():
    """One occupied coordinate: the file is named `A1`, never the degenerate `A1:A1`."""
    wb, s = _wb("Only")
    s["A1"] = "solo"
    _save(wb, "single_cell_sheet.xlsx")


def make_style_only_blank():
    """A cell whose WHOLE content is its fill: it states no value, occupies its coordinate anyway, and
    must come back through a pack as a `<c s=…>` with no `<v>` rather than leaving in silence."""
    wb, s = _wb("Blank")
    s["A1"] = "kept"
    s["B1"].fill = PatternFill(fill_type="solid", fgColor=BLANK_FILL)
    _save(wb, "style_only_blank.xlsx")


def make_hatch_and_diagonal_blanks():
    """Three blanks that all draw something in Excel, only ONE of which a rule can carry. `B1`'s solid
    fill has a `background-color`; `D4`'s gray125 hatch and `E5`'s diagonal edge have no declaration at
    all and are dropped and named. So only `B1` is occupancy: counting the other two occupies
    coordinates the export comes back WITHOUT, and the sheet's file grows from 2 TSV fields to 25 on
    the way in and shrinks back on the way out.

    `style_only_blank` holds the carried case alone, and every warn_* fixture puts its unspellable
    attribute on a cell that also holds a VALUE — so nothing in the corpus held a blank whose only look
    is one no rule can spell."""
    wb, s = _wb("Blanks")
    s["A1"] = "kept"
    s["B1"].fill = PatternFill(fill_type="solid", fgColor=BLANK_FILL)
    s["D4"].fill = PatternFill(patternType=HATCH_PATTERN)
    s["E5"].border = Border(diagonal=Side(style="thin", color=BLACK), diagonalUp=True)
    _save(wb, "hatch_and_diagonal_blanks.xlsx")


def make_empty_sheet():
    """No occupancy at all, so no range file — and nothing to pack, which the oracle's pack leg skips."""
    wb, _ = _wb("Empty")
    _save(wb, "empty_sheet.xlsx")


def make_warn_merged_region():
    """A merged region: a region fact CSS cannot carry, flattened and named."""
    wb, s = _wb("Sheet1")
    s["A1"] = "Title"
    s.merge_cells("A1:B1")
    _save(wb, "warn_merged_region.xlsx")


def make_warn_indent():
    """An indent level: expressible in CSS, deliberately not modelled, so it must be named."""
    wb, s = _wb("Sheet1")
    s["A1"] = "Indented"
    s["A1"].alignment = Alignment(indent=2)
    _save(wb, "warn_indent.xlsx")


def make_warn_dash_dot_border():
    """A `dashDot` edge: CSS `border-style` has no dash-dot, so it is approximated and named."""
    wb, s = _wb("Sheet1")
    s["A1"] = "Edge"
    s["A1"].border = Border(bottom=Side(style="dashDot", color=BLACK))
    _save(wb, "warn_dash_dot_border.xlsx")


def make_warn_underline_double():
    """`u="double"`: `TextDecoration` carries one underline, so the style is narrowed and named."""
    wb, s = _wb("Sheet1")
    s["A1"] = "Under"
    s["A1"].font = Font(name="Calibri", size=11, underline="double")
    _save(wb, "warn_underline_double.xlsx")


def make_warn_center_continuous():
    """`horizontal="centerContinuous"`: no `TextAlign` target, so it is dropped and named."""
    wb, s = _wb("Sheet1")
    s["A1"] = "Centered"
    s["A1"].alignment = Alignment(horizontal="centerContinuous")
    _save(wb, "warn_center_continuous.xlsx")


def _chart_book(title, headers, rows):
    """A sheet whose row 1 is the field names and whose body is the records, which is the shape a
    binding resolves to: a range keys on its FIRST ROW."""
    wb, s = _wb("Sheet1")
    s.append(headers)
    for row in rows:
        s.append(row)
    return wb, s


def make_bar_chart():
    """One series over `A1:B4`, the shape the whole read leg turns on: `<c:cat>` names column A and
    `<c:val>` column B, so neither reference is the binding and their bounding rectangle — extended up
    to the header row `<c:tx>` names — is. It must cross as a `bar` figure whose one `data.name` is
    `Sheet1!A1:B4` and whose `x` and `y` are the two header names.

    The chart is TITLED, because the title is what the figure is named for; an untitled one would fall
    back to the chart part and grade the fallback instead."""
    wb, s = _chart_book("Units", ("Region", "Units"), [("North", 12), ("South", 9), ("East", 15)])
    c = BarChart()
    c.title = "Units by region"
    c.add_data(Reference(s, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    c.set_categories(Reference(s, min_col=1, min_row=2, max_row=4))
    s.add_chart(c, "D2")
    _save(wb, "chart_bar_one_series.xlsx")


def make_two_series_line_chart():
    """Two series sharing one category column, which `chart_bar_one_series` cannot hold: ONE `<c:ser>`
    is one layer, so this must cross as a two-layer spec with two `data` objects — and the second
    series' value column is not the category column's neighbour, so it also grades that a binding is a
    bounding RECTANGLE rather than two adjacent columns."""
    wb, s = _chart_book(
        "Two", ("Month", "Alpha", "Beta"), [("Jan", 1, 2), ("Feb", 3, 4), ("Mar", 5, 6)]
    )
    c = LineChart()
    c.title = "Alpha against Beta"
    c.add_data(Reference(s, min_col=2, max_col=3, min_row=1, max_row=4), titles_from_data=True)
    c.set_categories(Reference(s, min_col=1, min_row=2, max_row=4))
    s.add_chart(c, "E2")
    _save(wb, "chart_line_two_series.xlsx")


def make_radar_chart_unsupported():
    """A radar, which has no Vega-Lite mark. Excel-to-Vega-Lite is TOTAL over the charts it admits, so
    this must yield NO figure and ONE named loss — never a silently different chart, and never a
    refusal: an unpack completes. Its series are otherwise ordinary, so what is graded is the MARK and
    nothing beside it."""
    wb, s = _chart_book("Radar", ("Axis", "Score"), [("A", 1), ("B", 2), ("C", 3)])
    c = RadarChart()
    c.title = "Scores"
    c.add_data(Reference(s, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    c.set_categories(Reference(s, min_col=1, min_row=2, max_row=4))
    s.add_chart(c, "D2")
    _save(wb, "chart_radar_unsupported.xlsx")


def make_unstyled_nothing_lost():
    """A workbook stating no style anywhere: the one fixture that must still print `nothing lost`."""
    wb, s = _wb("Plain")
    s["A1"], s["B1"] = "Name", "Score"
    s["A2"], s["B2"] = "Ada", 91
    s["A3"], s["B3"] = "Grace", 88
    _save(wb, "unstyled_nothing_lost.xlsx")


def main():
    make_styled_header_row()
    make_formatted_column()
    make_total_row_top_border()
    make_banded_body()
    make_body_8pt_vs_normal_11pt()
    make_normal_font_arial_9()
    make_sparse_blocks_normal_font_arial_9()
    make_column_and_row_default_style()
    make_column_run_default_style()
    make_theme1_color_noop()
    make_axis_width_and_height()
    make_title_over_table()
    make_stray_cell_sheet()
    make_single_cell_sheet()
    make_style_only_blank()
    make_hatch_and_diagonal_blanks()
    make_empty_sheet()
    make_warn_merged_region()
    make_warn_indent()
    make_warn_dash_dot_border()
    make_warn_underline_double()
    make_warn_center_continuous()
    make_unstyled_nothing_lost()
    make_bar_chart()
    make_two_series_line_chart()
    make_radar_chart_unsupported()


if __name__ == "__main__":
    main()
