# Concern: authors the decompose corpus's .xlsx fixtures — one structural concern per file | Non-concern: unpacking, grading, the frozen expectations | IO: () -> fixtures/*.xlsx

"""Regenerate the decomposition-conformance fixtures.

Run with any interpreter carrying openpyxl:

    python3 conformance/decompose/make_fixtures.py

The generated .xlsx ARE committed (small, LFS-tracked); nothing else here is generated. Provenance is
honest by construction: openpyxl (a third-party writer) authors every byte, never FSA1 — see
PROVENANCE.md.

Each fixture isolates ONE structural concern and is named for it, so a grader failure names the
concern rather than "the corpus". Every one is small enough that its expected block list is derivable
by hand from its own `s=` attributes; a fixture that is not is too big for this corpus.

The anchor shared with the expectations: a cell's SIGNATURE is its xf index — the `s=` attribute the
writer puts on `<c>`, absent where the cell states no style at all. An xf entry carries `numFmtId`
beside font, fill, border and alignment, so two cells share a signature only when they share the
whole look AND the display format.
"""

import datetime
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "fixtures")

TITLE_FONT = Font(bold=True, size=14)
CAPTION_FONT = Font(italic=True, size=9)
HEADER_FONT = Font(bold=True, size=11)
BODY_FONT = Font(size=10)
NOTE_HEADING_FONT = Font(bold=True, size=9)
NOTE_BODY_FONT = Font(size=9)
DATE_FORMAT = "yyyy-mm-dd"
COUNT_FORMAT = "#,##0"
FLUSH = Alignment(horizontal="left")
INDENTED = Alignment(horizontal="left", indent=2)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
BAND_EVEN_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
BAND_ODD_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")


def _save(wb, name):
    os.makedirs(OUT, exist_ok=True)
    wb.save(os.path.join(OUT, name))
    print(f"wrote fixtures/{name}")


def _wb(title):
    wb = openpyxl.Workbook()
    wb.active.title = title
    return wb, wb.active


def _write(sheet, coord, value, font=None, alignment=None, number_format=None, fill=None):
    cell = sheet[coord]
    cell.value = value
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format
    if fill is not None:
        cell.fill = fill


def make_title_caption_table():
    """A title line, a two-line caption and a bold-header table, stacked under one bounding box with
    no blank row anywhere: the whole sheet is 3x9, which the occupancy policy carries as ONE block.
    Four signatures, and the boundaries a reader reads are exactly the places one changes.

    The title takes TWO cells and the caption TWO rows deliberately. Two vertically adjacent
    single-row regions of the same width always coalesce under this policy — one row rule costs the
    same as the region term it replaces — so a one-cell title over a one-cell caption is one block,
    and the caption's own start row is then a boundary nothing writes. PROVENANCE.md records that
    re-authoring and the derivation behind it."""
    wb, s = _wb("Report")
    _write(s, "A1", "Quarterly Sales", TITLE_FONT)
    _write(s, "B1", "FY2026", TITLE_FONT)
    _write(s, "A2", "All figures in thousands of USD", CAPTION_FONT)
    _write(s, "A3", "Source: internal ledger", CAPTION_FONT)
    for col, head in zip("ABC", ("Region", "Units", "Revenue")):
        _write(s, f"{col}4", head, HEADER_FONT)
    body = [
        ("North", 12, 3400),
        ("South", 9, 2750),
        ("East", 15, 4100),
        ("West", 7, 1980),
        ("Central", 11, 3020),
    ]
    for i, row in enumerate(body, start=5):
        for col, value in zip("ABC", row):
            _write(s, f"{col}{i}", value, BODY_FONT)
    _save(wb, "title_caption_table.xlsx")


def make_table_then_footnotes():
    """A table, then three separately-headed footnote blocks, each set off by a SINGLE blank row —
    one strip, well inside the two the expansion may vault. So the leap reaches every one of those
    gaps and the cost model is the only thing holding them open: bridging a blank row suppresses the
    modal rule over the merged rectangle, and each remaining row then costs a rule of its own.

    The footnotes are one column wide against the table's three, so they also state the width
    mismatch that keeps the table itself off them."""
    wb, s = _wb("Notes")
    for col, head in zip("ABC", ("Item", "Qty", "Note")):
        _write(s, f"{col}1", head, HEADER_FONT)
    body = [
        ("Widget", 12, "stocked"),
        ("Gasket", 40, "stocked"),
        ("Flange", 7, "backorder"),
        ("Bracket", 22, "stocked"),
    ]
    for i, row in enumerate(body, start=2):
        for col, value in zip("ABC", row):
            _write(s, f"{col}{i}", value, BODY_FONT)
    notes = [
        (7, "Note 1", ["Backorder lead time is 14 days.", "Counted at the close of the month."]),
        (11, "Note 2", ["Quantities are whole units.", "Partial cases are rounded down."]),
        (15, "Note 3", ["Stocked means on hand today.", "The count excludes reserved stock."]),
    ]
    for row, heading, lines in notes:
        _write(s, f"A{row}", heading, NOTE_HEADING_FONT)
        for offset, line in enumerate(lines, start=1):
            _write(s, f"A{row + offset}", line, NOTE_BODY_FONT)
    _save(wb, "table_then_footnotes.xlsx")


def make_numfmt_only_series():
    """A workbook stating NO font, no fill and no border anywhere: a General-formatted preamble sits
    directly on top of a dated series, and the only thing separating them is the number format. The
    preamble's cells state no style at all, so the split exists only if an xf's `numFmtId` reaches
    the signature — read the look alone and every cell here states the same nothing, which the seed
    covers as one 2x11 rectangle and no cut can then reach.

    The two series columns carry DIFFERENT formats, so the pair also states the shape a column rule
    is for: one modal signature and one column that differs from it."""
    wb, s = _wb("Series")
    for i, (label, value) in enumerate(
        [("Report", "Widget sales"), ("Prepared", "Finance"), ("Basis", "Units")], start=1
    ):
        _write(s, f"A{i}", label)
        _write(s, f"B{i}", value)
    day = datetime.date(2026, 1, 5)
    counts = [120, 135, 98, 143, 151, 129, 164, 118]
    for i, count in enumerate(counts):
        _write(s, f"A{4 + i}", day + datetime.timedelta(days=7 * i), number_format=DATE_FORMAT)
        _write(s, f"B{4 + i}", count, number_format=COUNT_FORMAT)
    _save(wb, "numfmt_only_series.xlsx")


def make_contents_index():
    """A contents index whose two halves differ in NOTHING but their alignment: three chapter entries
    flush left, then three subsection entries indented under the last of them. Same font, same fill,
    same border, same number format, no blank row between them — an xf entry carries an `alignment`
    child, so the two halves state two signatures and the index is two regions.

    Both halves are three rows deep on purpose: a merge that costs one row rule per minority row pays
    for itself while one side is a single row, and stops paying at two."""
    wb, s = _wb("Contents")
    chapters = [("Introduction", 1), ("Methods", 4), ("Results", 9)]
    subsections = [("Sampling", 10), ("Instruments", 12), ("Analysis", 14)]
    for i, (title, page) in enumerate(chapters, start=1):
        _write(s, f"A{i}", title, alignment=FLUSH)
        _write(s, f"B{i}", page, alignment=FLUSH)
    for i, (title, page) in enumerate(subsections, start=4):
        _write(s, f"A{i}", title, alignment=INDENTED)
        _write(s, f"B{i}", page, alignment=INDENTED)
    _save(wb, "contents_index.xlsx")


def make_banded_report():
    """A bold header over an eight-row zebra body, every cell carrying a value: the ONLY thing that
    moves down this sheet is the body's fill, alternating white on the even rows and grey on the odd
    ones. Three signatures over 27 cells — header, even band, odd band — and every boundary a reader
    could read here is a row boundary, with no blank row and no width change anywhere to state one.

    The body is eight rows so that each band holds four. A merge that costs one row rule per minority
    row pays for itself while one side is a single row and stops paying at two, so four a side puts
    the two bands well past any accounting under which they coalesce: the alternation is stated, not
    merely present. The header takes a fill of its own rather than a bold font alone, so the top
    boundary is a fill change as well as a font change and cannot be read as the first band."""
    wb, s = _wb("Report")
    for col, head in zip("ABC", ("Region", "Units", "Revenue")):
        _write(s, f"{col}1", head, HEADER_FONT, fill=HEADER_FILL)
    body = [
        ("North", 128, 4210),
        ("South", 96, 3155),
        ("East", 143, 4680),
        ("West", 74, 2490),
        ("Central", 111, 3620),
        ("Northeast", 88, 2940),
        ("Southwest", 132, 4305),
        ("Midwest", 105, 3480),
    ]
    for i, row in enumerate(body, start=2):
        band = BAND_EVEN_FILL if i % 2 == 0 else BAND_ODD_FILL
        for col, value in zip("ABC", row):
            _write(s, f"{col}{i}", value, BODY_FONT, fill=band)
    _save(wb, "banded_report.xlsx")


def make_banded_subtotals():
    """The same zebra body cut three times by a bold subtotal row on the header's own fill, and the
    banding runs on GLOBAL row parity straight through: row 5 resumes the alternation the subtotal on
    row 4 interrupted, never restarting per group. Thirty cells, three signatures — even band, odd
    band, and the header/subtotal fill that lands on rows 1, 4, 7 and 10 — so the subtotal rows are
    the only interruption in an otherwise uniform stripe.

    Global parity is what the fixture is for, and it states two facts a per-group restart would hide.
    The groups are not identical: rows 2-3 and 8-9 read white-then-grey while rows 5-6 read
    grey-then-white, so no repeated three-row unit can cover the body. And because each subtotal row
    consumes one parity slot, the band directly above it matches the band directly below — rows 3 and
    5 are both grey, rows 6 and 8 both white — so a region grown across a subtotal row lands back on
    the fill it left, and the subtotal is a one-row hole inside a look that resumes unchanged."""
    wb, s = _wb("Sales")
    for col, head in zip("ABC", ("Account", "Units", "Revenue")):
        _write(s, f"{col}1", head, HEADER_FONT, fill=HEADER_FILL)
    body = [
        ("North", 120, 4800),
        ("South", 95, 3800),
        ("Subtotal Q1", 215, 8600),
        ("East", 140, 5600),
        ("West", 88, 3520),
        ("Subtotal Q2", 228, 9120),
        ("Central", 105, 4200),
        ("Mountain", 76, 3040),
        ("Subtotal Q3", 181, 7240),
    ]
    subtotal_rows = {4, 7, 10}
    for i, row in enumerate(body, start=2):
        if i in subtotal_rows:
            font, fill = HEADER_FONT, HEADER_FILL
        else:
            font = BODY_FONT
            fill = BAND_EVEN_FILL if i % 2 == 0 else BAND_ODD_FILL
        for col, value in zip("ABC", row):
            _write(s, f"{col}{i}", value, font, fill=fill)
    _save(wb, "banded_subtotals.xlsx")


def main():
    make_title_caption_table()
    make_table_then_footnotes()
    make_numfmt_only_series()
    make_contents_index()
    make_banded_report()
    make_banded_subtotals()


if __name__ == "__main__":
    main()
