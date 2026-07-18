# Concern: the ENG6 per-formula differential harness — for each corpus case, compute a REFERENCE value with the `formulas` Excel lib (openpyxl-built .xlsx) and a CHARLIE value via `charlie-cli eval` over an equivalent filesystem workbook, classify each MATCH / DIVERGE / EXCLUDED (declared lib-gap), print a parity table + summary, and exit non-zero on any charlie-defect divergence | Non-concern: WHAT charlie computes or WHY (that is charlie-model/charlie-ast); WHETHER the `formulas` lib is itself right for a case (a human triages a DIVERGE and records genuine lib limitations in KNOWN-LIB-GAPS.md, then flags the case `lib_gap` to exclude it — this harness never edits charlie to chase a wrong reference) | IO: (in: corpus/*.json cases, the charlie-cli binary, a temp workdir for generated .xlsx + charlie workbooks) -> a parity table + match/total summary on stdout, exit 0 iff no charlie-defect divergence
"""ENG6 differential oracle: grade charlie cell-for-cell against the `formulas` Excel reference.

Run via ``run.sh`` (which activates the venv). Each corpus case is
``{name, category, inputs: {A1: literal, ...}, formula: "=...", tol?, lib_gap?, lib_gap_reason?}``.
The reference is computed live by the `formulas` library over an openpyxl-built workbook; the
charlie value comes from ``charlie-cli eval``. Comparison is driven by the *reference's* value type
(number within tol, else string-exact), so a text result like ``TEXT(..,"0.00") -> "1234.50"`` is
compared as text, not silently reparsed as a number.
"""

import json
import numbers
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import numpy as np
import openpyxl

# Silence the `formulas` lib's tqdm progress bars + chatty logging before importing it.
import logging

logging.disable(logging.CRITICAL)
os.environ.setdefault("TQDM_DISABLE", "1")
import formulas  # noqa: E402

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parents[1]
CORPUS_DIR = HERE / "corpus"
DEFAULT_CLI = REPO_ROOT / "target" / "debug" / "charlie-cli"
TAB = "S"  # single tab, in both the xlsx sheet and the charlie tab folder

# The nine canonical Excel error spellings — used to classify a reference value as an error kind.
EXCEL_ERRORS = {
    "#DIV/0!",
    "#N/A",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#NULL!",
    "#SPILL!",
    "#CALC!",
}

REL_TOL = 1e-9  # relative floor added to every numeric comparison alongside the per-case abs tol.


def load_corpus():
    """Load and flatten every case in ``corpus/*.json``, sorted by (category, name)."""
    cases = []
    for path in sorted(CORPUS_DIR.glob("*.json")):
        with open(path, encoding="utf-8") as handle:
            batch = json.load(handle)
        for case in batch:
            case.setdefault("category", path.stem)
            case.setdefault("inputs", {})
            case.setdefault("tol", 0.0)
            cases.append(case)
    cases.sort(key=lambda c: (c["category"], c["name"]))
    return cases


def _unwrap(value):
    """Reduce a `formulas` result to a scalar: a 1-element ndarray/Array -> its element; else as-is."""
    if isinstance(value, np.ndarray):
        flat = value.ravel()
        if flat.size == 1:
            return flat[0]
        return value  # a genuine multi-cell array (no corpus case relies on this today)
    return value


def reference_value(inputs, formula, workdir):
    """Build a tiny .xlsx (inputs in their cells, formula in Z100) and compute it with `formulas`.

    Returns the raw scalar reference value (a number, str, bool, or an XlError-like error object).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TAB
    for cell, literal in inputs.items():
        ws[cell] = literal
    ws["Z100"] = formula if formula.startswith("=") else "=" + formula
    xlsx_path = workdir / "ref.xlsx"
    wb.save(xlsx_path)
    model = formulas.ExcelModel().loads(str(xlsx_path)).finish()
    solution = model.calculate()
    key = next(k for k in solution if k.upper().endswith("!Z100"))
    return _unwrap(solution[key].value)


def build_charlie_workbook(inputs, workdir):
    """Lay each input literal into a single-cell grid file under a ``S/`` tab folder (grid-only)."""
    wb_dir = workdir / "cwb"
    tab_dir = wb_dir / TAB
    tab_dir.mkdir(parents=True, exist_ok=True)
    if not inputs:
        # A tab needs at least one file to be a tab; a formula over no inputs still needs the folder.
        (tab_dir / "A1").write_text("", encoding="utf-8")
    for cell, literal in inputs.items():
        (tab_dir / cell).write_text(_grid_literal(literal), encoding="utf-8")
    return wb_dir


def _grid_literal(literal):
    """Render a corpus literal as charlie grid content (a bare number/text cell — no annotation)."""
    if isinstance(literal, bool):
        return "TRUE" if literal else "FALSE"
    return str(literal)


def charlie_value(cli, wb_dir, formula, tab=TAB):
    """Run ``charlie-cli eval`` and return ``(value_str_or_None, exit_code, note)``.

    ``value_str`` is the string in the JSON envelope's ``data.value`` (present for both a plain value
    and an error-valued result like ``#DIV/0!``); ``None`` means a parse refusal (only diagnostics).
    ``tab`` selects the sheet unqualified refs bind to — the per-formula oracle uses the single ``S``
    tab; the whole-workbook oracle passes each cell's own sheet.
    """
    proc = subprocess.run(
        [str(cli), "eval", str(wb_dir), "--tab", tab, "--formula", formula, "--format", "json"],
        capture_output=True,
        text=True,
    )
    try:
        envelope = json.loads(proc.stdout)
    except json.JSONDecodeError:
        return None, proc.returncode, f"non-JSON stdout: {proc.stdout.strip()[:80]}"
    data = envelope.get("data") or {}
    if "value" in data:
        return data["value"], proc.returncode, ""
    diags = data.get("diagnostics") or []
    msg = diags[0].get("message", "parse refusal") if diags else "parse refusal"
    return None, proc.returncode, msg


def classify_ref(value):
    """Classify a reference scalar into (kind, canonical_string). kind in {error,bool,num,text}."""
    text = str(value)
    if text in EXCEL_ERRORS:
        return "error", text
    if isinstance(value, (bool, np.bool_)):
        return "bool", "TRUE" if bool(value) else "FALSE"
    if isinstance(value, numbers.Number):
        return "num", value
    if isinstance(value, str):
        return "text", value
    # Anything else (e.g. an unexpected lib sentinel) is treated as its string form, exact-compared.
    return "text", text


def compare(case, ref_value, charlie_str):
    """Compare charlie's raw output string against the reference, driven by the reference's kind.

    Returns ``(is_match, ref_display, kind, detail)``.
    """
    kind, canonical = classify_ref(ref_value)
    if kind == "error":
        ok = charlie_str == canonical
        return ok, canonical, kind, "" if ok else "charlie did not reproduce the error value"
    if kind == "bool":
        ok = charlie_str == canonical
        return ok, canonical, kind, "" if ok else "boolean mismatch"
    if kind == "num":
        if charlie_str is None:
            return False, _fmt_num(canonical), kind, "charlie produced no value (refusal)"
        try:
            got = float(charlie_str)
        except ValueError:
            return False, _fmt_num(canonical), kind, f"charlie value {charlie_str!r} is not numeric"
        tol = float(case.get("tol", 0.0))
        ok = abs(got - float(canonical)) <= tol + REL_TOL * abs(float(canonical))
        return ok, _fmt_num(canonical), kind, "" if ok else "numeric divergence beyond tolerance"
    # text
    ok = charlie_str == canonical
    return ok, canonical, kind, "" if ok else "text mismatch"


def _fmt_num(value):
    """Compact numeric display for the parity table."""
    f = float(value)
    if f == int(f) and abs(f) < 1e15:
        return str(int(f))
    return repr(f)


def _clip(text, width):
    text = "" if text is None else str(text)
    return text if len(text) <= width else text[: width - 1] + "…"


def run():
    cli = Path(os.environ.get("CHARLIE_CLI", DEFAULT_CLI))
    if not cli.exists():
        print(f"charlie-cli not found at {cli} (build it: cargo build -p charlie-cli)", file=sys.stderr)
        return 2
    cases = load_corpus()
    if not cases:
        print("no corpus cases found under corpus/*.json", file=sys.stderr)
        return 2

    artifacts = HERE / ".artifacts"
    artifacts.mkdir(exist_ok=True)  # import-safe: the temp parent must exist before mkdtemp, not only under __main__.
    work_root = Path(tempfile.mkdtemp(prefix="xl-oracle-", dir=str(artifacts)))
    rows = []
    defects = 0
    matched = 0
    graded = 0
    try:
        for case in cases:
            workdir = work_root / case["name"]
            workdir.mkdir(parents=True, exist_ok=True)
            lib_gap = bool(case.get("lib_gap"))

            ref_err = None
            try:
                ref = reference_value(case["inputs"], case["formula"], workdir)
            except Exception as exc:  # a lib that cannot even evaluate the case is a reference failure
                ref = None
                ref_err = f"{type(exc).__name__}: {exc}".splitlines()[0][:60]

            wb_dir = build_charlie_workbook(case["inputs"], workdir)
            ch_value, ch_exit, ch_note = charlie_value(cli, wb_dir, case["formula"])

            if ref_err is not None:
                verdict = "EXCLUDED" if lib_gap else "DIVERGE"
                rows.append((case, verdict, f"<ref err: {ref_err}>", ch_value, ch_note or "reference lib failed"))
                if not lib_gap:
                    defects += 1
                continue

            is_match, ref_disp, kind, detail = compare(case, ref, ch_value)
            if lib_gap:
                rows.append((case, "EXCLUDED", ref_disp, ch_value, case.get("lib_gap_reason", "")))
                continue
            graded += 1
            if is_match:
                matched += 1
                rows.append((case, "MATCH", ref_disp, ch_value, kind))
            else:
                defects += 1
                rows.append((case, "DIVERGE", ref_disp, ch_value, detail))
    finally:
        shutil.rmtree(work_root, ignore_errors=True)

    _print_table(rows)
    excluded = sum(1 for _, v, *_ in rows if v == "EXCLUDED")
    print()
    print(f"parity: {matched}/{graded} graded cases match  "
          f"(excluded lib-gaps: {excluded}; total corpus: {len(cases)})")
    if defects:
        print(f"DIVERGENCES (charlie defects): {defects}")
    else:
        print("no charlie-defect divergences")
    return 1 if defects else 0


def _print_table(rows):
    headers = ("category", "case", "verdict", "reference", "charlie", "detail")
    widths = (12, 20, 8, 20, 20, 40)
    line = "  ".join(h.ljust(w) for h, w in zip(headers, widths))
    print(line)
    print("  ".join("-" * w for w in widths))
    for case, verdict, ref_disp, ch_value, detail in rows:
        cells = (
            _clip(case["category"], widths[0]).ljust(widths[0]),
            _clip(case["name"], widths[1]).ljust(widths[1]),
            verdict.ljust(widths[2]),
            _clip(ref_disp, widths[3]).ljust(widths[3]),
            _clip(ch_value, widths[4]).ljust(widths[4]),
            _clip(detail, widths[5]),
        )
        print("  ".join(cells))


if __name__ == "__main__":
    sys.exit(run())
