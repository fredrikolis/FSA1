# Concern: AUTHOR the committed real-file .xlsx corpus for the WHOLE-WORKBOOK differential oracle (DoD milestone 3) — emit a few small but realistic multi-formula Excel workbooks (a mini P&L income statement, a loan amortization schedule, and a price lookup table) via openpyxl, each mixing literals with a web of formulas (SUM/subtraction/percentage, PMT/IPMT/PPMT and a running-balance chain, VLOOKUP/INDEX-MATCH/IF) so the oracle can import->compute->diff charlie against the `formulas` reference on genuine files, not single-formula probes | Non-concern: COMPUTING or DIFFING the values (workbook_oracle.py owns the import->eval->compare), the per-formula JSON corpus (corpus/*.json), and the reference stack pin (requirements.txt) | IO: () -> writes corpus_workbooks/*.xlsx next to this script
"""Regenerate the committed real-file .xlsx workbook corpus for workbook_oracle.py.

Run via the oracle venv (created by run.sh):

    conformance/xl-oracle/.venv/bin/python conformance/xl-oracle/make_workbooks.py

The generated .xlsx ARE committed (small); the venv is gitignored. Formulas are written in Excel-A1
so both the `formulas` reference and charlie (after `charlie-cli import`) evaluate the same source.
"""

import os

import openpyxl

HERE = os.path.dirname(__file__)
OUT = os.path.join(HERE, "corpus_workbooks")


def write(wb, name):
    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, name)
    wb.save(path)
    print("wrote", os.path.normpath(path))


def make_pnl():
    """A mini profit-and-loss: revenue/COGS/opex lines, subtotals, a tax %, and net margin — a web of
    SUM, subtraction, multiplication, and division over a two-column (Q1/Q2) statement, plus a Summary
    sheet that references the P&L cross-sheet."""
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "PnL"
    # Column A: labels; B: Q1; C: Q2. Rows chain into subtotals.
    rows = [
        ("Product revenue", 120000, 138000),
        ("Service revenue", 45000, 52000),
    ]
    s["A1"] = rows[0][0]; s["B1"] = rows[0][1]; s["C1"] = rows[0][2]
    s["A2"] = rows[1][0]; s["B2"] = rows[1][1]; s["C2"] = rows[1][2]
    s["A3"] = "Total revenue"; s["B3"] = "=SUM(B1:B2)"; s["C3"] = "=SUM(C1:C2)"
    s["A4"] = "COGS"; s["B4"] = 66000; s["C4"] = 73000
    s["A5"] = "Gross profit"; s["B5"] = "=B3-B4"; s["C5"] = "=C3-C4"
    s["A6"] = "Gross margin"; s["B6"] = "=B5/B3"; s["C6"] = "=C5/C3"
    s["A7"] = "Operating expenses"; s["B7"] = 38000; s["C7"] = 41000
    s["A8"] = "Operating income"; s["B8"] = "=B5-B7"; s["C8"] = "=C5-C7"
    s["A9"] = "Tax @ 21%"; s["B9"] = "=ROUND(B8*0.21,2)"; s["C9"] = "=ROUND(C8*0.21,2)"
    s["A10"] = "Net income"; s["B10"] = "=B8-B9"; s["C10"] = "=C8-C9"
    s["A11"] = "Full-year net"; s["B11"] = "=B10+C10"
    summary = wb.create_sheet("Summary")
    summary["A1"] = "FY net income"; summary["B1"] = "=PnL!B11"
    summary["A2"] = "H1 revenue"; summary["B2"] = "=PnL!B3+PnL!C3"
    summary["A3"] = "Avg gross margin"; summary["B3"] = "=AVERAGE(PnL!B6:C6)"
    write(wb, "pnl.xlsx")


def make_amortization():
    """A loan amortization schedule: a monthly PMT, then a per-period chain of interest / principal /
    ending-balance formulas, plus IPMT/PPMT cross-checks and a total-interest SUM."""
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Loan"
    # Inputs.
    s["A1"] = "Principal"; s["B1"] = 20000
    s["A2"] = "Annual rate"; s["B2"] = 0.06
    s["A3"] = "Months"; s["B3"] = 6
    s["A4"] = "Monthly payment"; s["B4"] = "=PMT(B2/12,B3,-B1)"
    # Schedule header row 6; periods 1..6 in rows 7..12.
    s["A6"] = "Period"; s["B6"] = "Interest"; s["C6"] = "Principal"; s["D6"] = "=B1"  # opening balance
    s["A7"] = 1
    s["B7"] = "=D6*$B$2/12"        # period-1 interest off the opening balance in D6
    s["C7"] = "=$B$4-B7"
    s["D7"] = "=D6-C7"
    for r in range(8, 13):
        s[f"A{r}"] = f"=A{r-1}+1"
        s[f"B{r}"] = f"=D{r-1}*$B$2/12"
        s[f"C{r}"] = f"=$B$4-B{r}"
        s[f"D{r}"] = f"=D{r-1}-C{r}"
    s["A14"] = "Total interest"; s["B14"] = "=SUM(B7:B12)"
    s["A15"] = "IPMT p1"; s["B15"] = "=IPMT(B2/12,1,B3,-B1)"
    s["A16"] = "PPMT p1"; s["B16"] = "=PPMT(B2/12,1,B3,-B1)"
    write(wb, "amortization.xlsx")


def make_lookup():
    """A price lookup table with VLOOKUP, INDEX/MATCH, HLOOKUP, IF tiers, and a SUMIF rollup — the
    canonical 'join a fact against a dimension table' shape."""
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Catalog"
    # Price table: A=SKU, B=price, C=category (rows 1..4).
    table = [("A100", 9.5, "widget"), ("A200", 14.0, "gadget"),
             ("A300", 3.25, "widget"), ("A400", 22.0, "gadget")]
    for i, (sku, price, cat) in enumerate(table, start=1):
        s[f"A{i}"] = sku; s[f"B{i}"] = price; s[f"C{i}"] = cat
    # Order lines: E=SKU, F=qty (rows 1..3).
    orders = [("A200", 3), ("A300", 10), ("A100", 2)]
    for i, (sku, qty) in enumerate(orders, start=1):
        s[f"E{i}"] = sku; s[f"F{i}"] = qty
    # G: unit price via VLOOKUP; H: line total; I: category via INDEX/MATCH; J: bulk flag via IF.
    for i in range(1, 4):
        s[f"G{i}"] = f"=VLOOKUP(E{i},$A$1:$C$4,2,FALSE)"
        s[f"H{i}"] = f"=G{i}*F{i}"
        s[f"I{i}"] = f"=INDEX($C$1:$C$4,MATCH(E{i},$A$1:$A$4,0))"
        s[f"J{i}"] = f'=IF(F{i}>=5,"bulk","std")'
    s["H5"] = "=SUM(H1:H3)"                          # order total
    s["H6"] = '=SUMIF(I1:I3,"widget",H1:H3)'         # widget revenue
    s["H7"] = "=HLOOKUP(9.5,B1:B4,1,FALSE)"          # numeric HLOOKUP over the price column
    write(wb, "lookup.xlsx")


def make_forging():
    """A reference-forging workbook (ENG6): dynamic OFFSET ranges driven by COUNT, INDIRECT resolving
    A1 text (bare and cross-sheet), and OFFSET shifts — the 'dynamic named range' + 'dynamic
    addressing' workhorses. charlie source-rewrites each forger to a static reference before evaluation;
    the `formulas` reference computes them natively, so both sides agree on the forged value. NESTED
    forging (a forger whose argument forges) is a deliberate charlie divergence (a located #REF!,
    restricted v1) and so is deliberately NOT authored here (the parity corpus grades supported cases;
    the refusal is pinned by the charlie-model forge fitness tests)."""
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Forge"
    # A1..A4 a data column; some formulas forge ranges/refs over it.
    for i, v in enumerate([10, 20, 30, 40], start=1):
        s[f"A{i}"] = v
    s["C1"] = "=SUM(OFFSET($A$1,0,0,3,1))"                     # static -> SUM(A1:A3) = 60
    s["C2"] = "=SUM(OFFSET($A$1,0,0,COUNT($A$1:$A$4),1))"     # dynamic height -> SUM(A1:A4) = 100
    s["C3"] = "=OFFSET($A$1,1,0)"                              # single-cell shift -> A2 = 20
    s["C4"] = '=INDIRECT("A"&2)'                               # A1 text built by concat -> A2 = 20
    s["C5"] = "=SUM(OFFSET($A$1,1,0,2,1))"                     # shifted 2-tall range -> SUM(A2:A3) = 50
    # A cross-sheet INDIRECT: resolve a reference on another sheet by text.
    d = wb.create_sheet("Data")
    d["B2"] = 77
    s["C6"] = '=INDIRECT("Data!B2")'                          # cross-sheet forge -> 77
    write(wb, "forging.xlsx")


if __name__ == "__main__":
    make_pnl()
    make_amortization()
    make_lookup()
    make_forging()
